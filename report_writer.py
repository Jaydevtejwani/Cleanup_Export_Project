import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font


def write_reports(all_results, output_file):
    """
    Step 1: Write all rule sheets + summary using pandas
    """

    summary_data = []

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

        for rule_name, df in all_results.items():

            sheet_name = rule_name[:31]  # Excel limit

            if df is None or df.empty:
                count = 0
            else:
                count = len(df)

            summary_data.append({
                "Rule Name": rule_name,
                "Count": count,
                "Sheet Name": sheet_name
            })

            if df is not None and not df.empty:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Create summary sheet
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name="SUMMARY", index=False)

    print("[INFO] Base Excel written successfully")


# ---------------------------------------------------
# STEP 2: Add hyperlinks + Back navigation
# ---------------------------------------------------
def add_back_links(output_file):

    wb = load_workbook(output_file)

    # ---------------------------
    # SUMMARY SHEET HYPERLINKS
    # ---------------------------
    summary_ws = wb["SUMMARY"]

    for row in range(2, summary_ws.max_row + 1):

        sheet_name = summary_ws[f"C{row}"].value
        rule_name_cell = summary_ws[f"A{row}"]

        if sheet_name:
            rule_name_cell.hyperlink = f"#{sheet_name}!A1"
            rule_name_cell.style = "Hyperlink"

    # ---------------------------
    # EACH SHEET BACK LINK
    # ---------------------------
    for sheet in wb.sheetnames:

        if sheet == "SUMMARY":
            continue

        ws = wb[sheet]

        # Shift data down to avoid header overlap
        ws.insert_rows(1)

        ws["A1"] = "⬅ Back to Summary"
        ws["A1"].hyperlink = "#SUMMARY!A1"
        ws["A1"].style = "Hyperlink"

        # Optional: freeze header row
        ws.freeze_panes = "A2"

    wb.save(output_file)

    print("[INFO] Hyperlinks + Back navigation added successfully")
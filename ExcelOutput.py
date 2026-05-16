import pandas as pd
import os
from glob import glob
from openpyxl import load_workbook
from openpyxl.styles import Font

folder_path = r"C:\Users\ADMIN\Desktop\Cleanup_Extraction Project\Project\Files"

excel_files = glob(os.path.join(folder_path, "*.xlsx"))

excel_files = [
    file for file in excel_files
    if os.path.basename(file) != "Master_File.xlsx"
]

output_file = os.path.join(folder_path, "Master_File.xlsx")
summary_data = []
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

    for file in excel_files:

        try:
            
            df = pd.read_excel(file)
            df = df.dropna(how='all')
            empty_columns = df.columns[df.isna().all()].tolist()
            if empty_columns:
                print(f"Skipped: {os.path.basename(file)}")
                print(f"Empty Columns: {empty_columns}")
                continue
            sheet_name = os.path.splitext(os.path.basename(file))[0][:31]

            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                startrow=2
            )

        
            summary_data.append([sheet_name, len(df)])

            print(f"Merged: {sheet_name}")

        except Exception as e:
            print(f"Error in {file}: {e}")

    
    summary_df = pd.DataFrame(
        summary_data,
        columns=["Checks Name", "Total Count"]
    )

    summary_df.to_excel(
        writer,
        sheet_name="Summary",
        index=False
    )


wb = load_workbook(output_file)

blue_font = Font(color="0000FF", underline="single")

summary_ws = wb["Summary"]

for row in range(2, summary_ws.max_row + 1):

    sheet_name = summary_ws[f"A{row}"].value

    summary_ws[f"A{row}"].hyperlink = f"#'{sheet_name}'!A1"
    summary_ws[f"A{row}"].font = blue_font

for sheet in wb.sheetnames:

    if sheet != "Summary":

        ws = wb[sheet]
        ws["A1"] = "Back To Summary"
        ws["A1"].hyperlink = "#'Summary'!A1"
        ws["A1"].font = blue_font
wb.save(output_file)

print("\n Master file created successfully!")